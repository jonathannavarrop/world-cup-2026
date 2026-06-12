#!/usr/bin/env python3
"""
Extrae las predicciones de cada jugador desde su Excel (hoja WORLDCUP)
y las normaliza en data/predictions.json.

Solo hay que ejecutarlo cuando alguien cambie sus pronósticos.
Uso:  python engine/extract_predictions.py /ruta/a/los/excels
"""
import json, re, sys, glob, os
from openpyxl import load_workbook
sys.path.insert(0, os.path.dirname(__file__))
from _names import en

# Fichero -> alias del jugador
PLAYERS = {
    "CHURU":    "Excel-Mundial-2026_CHURU.xlsx",
    "JONY":     "Excel-Mundial-2026_JONY.xlsx",
    "CHAMP":    "Excel-Mundial-2026-CHAMP.xlsx",
    "NEGREIRA": "Excel-Mundial-2026-NEGREIRA.xlsx",
    "PEP":      "Excel-Mundial-2026-PEP.xlsx",
}

# Filas de cada ronda eliminatoria en la plantilla (columnas AA=27 local, AF=32 visitante)
KO_ROWS = {
    "R32":   range(101, 117),
    "R16":   range(120, 128),
    "QF":    range(131, 135),
    "SF":    [138, 139],
    "Final": [147],
}

def extract_one(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["WORLDCUP"]

    # --- Fase de grupos: marcador exacto por partido (1..72) ---
    group = {}
    for r in range(4, 98):
        num = ws.cell(r, 34).value          # AH = nº de partido
        home, away = ws.cell(r, 27).value, ws.cell(r, 32).value
        gh, ga = ws.cell(r, 29).value, ws.cell(r, 30).value
        if isinstance(num, (int, float)) and 1 <= num <= 72 and home and away:
            group[int(num)] = {
                "home": en(home), "away": en(away),
                "gh": None if gh is None else int(gh),
                "ga": None if ga is None else int(ga),
            }

    # --- Clasificación pronosticada de cada grupo (posición 1A..4L) ---
    # Sirve para el bonus de "misma posición" (+30) y para el premio LOBBY.
    standings = {}
    for r in range(4, 98):
        code = ws.cell(r, 35).value         # AI = código de posición, p.ej. "1A"
        if code and re.match(r"^[1-4][A-L]$", str(code).strip()):
            code = str(code).strip()
            standings[code] = {
                "team": en(ws.cell(r, 38).value),    # AL
                "pts":  ws.cell(r, 39).value,        # AM
                "gf":   ws.cell(r, 44).value,        # AR
                "gd":   ws.cell(r, 46).value,        # AT
            }

    # --- Bracket: equipos que el jugador hace llegar a cada ronda ---
    advance = {}
    for rnd, rows in KO_ROWS.items():
        teams = []
        for r in rows:
            for col in (27, 32):
                v = ws.cell(r, col).value
                if v and not str(v).startswith(("W", "L")):  # ignora refs sin resolver
                    teams.append(en(v))
        advance[rnd] = sorted(set(teams))
    advance["Champion"] = [en(ws.cell(150, 27).value)] if ws.cell(150, 27).value and not str(ws.cell(150, 27).value).startswith(("W", "L")) else []

    # --- Premios extra (todavía sin rellenar en los Excel -> se editan a mano) ---
    prizes = {
        "roberswingger": None,           # 1 jugador (máximo goleador)
        "champ": None,                   # 1 jugador (mejor jugador)
        "paraguaya_chupona": None,       # 1 selección (la más agresiva)
        "pinwis": [None, None],          # 2 jugadores
        "lobby": None,                   # 1 selección (la peor)
        "columna_churu": None,           # 1 portero
    }

    return {"group": group, "standings": standings, "advance": advance, "prizes": prizes}


def main():
    base = sys.argv[1] if len(sys.argv) > 1 else "."
    out = {}
    for alias, fname in PLAYERS.items():
        path = os.path.join(base, fname)
        if not os.path.exists(path):
            print(f"  ! falta {fname}, lo salto")
            continue
        out[alias] = extract_one(path)
        g = sum(1 for v in out[alias]["group"].values() if v["gh"] is not None)
        adv = {k: len(v) for k, v in out[alias]["advance"].items()}
        print(f"  {alias:9} grupos={g}/72  bracket={adv}")

    os.makedirs("data", exist_ok=True)
    with open("data/predictions.json", "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print("-> data/predictions.json")


if __name__ == "__main__":
    main()
